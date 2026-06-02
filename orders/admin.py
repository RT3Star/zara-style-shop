from django.contrib import admin
from django.utils.safestring import mark_safe
from import_export.admin import ExportActionModelAdmin
from .models import Order, OrderItem
from .resources import OrderResource, OrderItemResource


class OrderItemInline(admin.TabularInline):
    model = OrderItem
    extra = 0
    readonly_fields = (
        "product",
        "product_name",
        "product_price",
        "quantity",
        "size",
        "color",
    )
    can_delete = False


@admin.register(Order)
class OrderAdmin(ExportActionModelAdmin, admin.ModelAdmin):
    resource_class = OrderResource
    list_display = (
        "id",
        "user",
        "get_full_name",
        "total_price",
        "status_badge",
        "is_paid",
        "created_at",
    )
    list_filter = (
        "status",
        "is_paid",
        "payment_method",
        "delivery_method",
        "created_at",
    )
    search_fields = (
        "id",
        "first_name",
        "last_name",
        "email",
        "phone",
        "user__username",
    )
    readonly_fields = ("created_at", "updated_at", "subtotal")
    list_editable = ("is_paid",)

    actions = ["export_as_csv", "export_as_excel"]

    fieldsets = (
        (
            "Інформація про покупця",
            {"fields": ("user", "first_name", "last_name", "email", "phone")},
        ),
        ("Адреса доставки", {"fields": ("address", "city", "postal_code")}),
        (
            "Доставка",
            {"fields": ("delivery_method", "delivery_price", "delivery_tracking")},
        ),
        ("Оплата", {"fields": ("payment_method", "is_paid", "paid_at")}),
        ("Фінанси", {"fields": ("subtotal", "total_price")}),
        ("Статус", {"fields": ("status", "admin_comment")}),
        (
            "Дати",
            {
                "fields": ("created_at", "updated_at"),
                "classes": ("collapse",),
            },
        ),
    )

    inlines = [OrderItemInline]

    def get_full_name(self, obj):
        return obj.get_full_name()

    get_full_name.short_description = "ПІБ"

    def status_badge(self, obj):
        colors = {
            "pending": "warning",
            "processing": "info",
            "shipped": "primary",
            "delivered": "success",
            "cancelled": "danger",
            "refunded": "secondary",
        }
        return mark_safe(
            f'<span class="badge bg-{colors.get(obj.status, "secondary")}">{obj.get_status_display()}</span>'
        )

    status_badge.short_description = "Статус"

    def export_as_csv(self, request, queryset):
        """Експорт у CSV"""
        import csv
        from django.http import HttpResponse

        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="orders.csv"'

        writer = csv.writer(response)
        writer.writerow(["ID", "Клієнт", "Email", "Телефон", "Сума", "Статус", "Дата"])

        for order in queryset:
            writer.writerow(
                [
                    order.id,
                    order.get_full_name(),
                    order.email,
                    order.phone,
                    order.total_price,
                    order.get_status_display(),
                    order.created_at.strftime("%d.%m.%Y %H:%M"),
                ]
            )

        return response

    export_as_csv.short_description = "Експортувати вибрані у CSV"

    def export_as_excel(self, request, queryset):
        """Експорт у Excel"""
        import openpyxl
        from django.http import HttpResponse

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="orders.xlsx"'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Замовлення"

        # Заголовки
        headers = ["ID", "Клієнт", "Email", "Телефон", "Сума", "Статус", "Дата"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
            ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)

        # Дані
        for row, order in enumerate(queryset, 2):
            ws.cell(row=row, column=1, value=order.id)
            ws.cell(row=row, column=2, value=order.get_full_name())
            ws.cell(row=row, column=3, value=order.email)
            ws.cell(row=row, column=4, value=order.phone)
            ws.cell(row=row, column=5, value=float(order.total_price))
            ws.cell(row=row, column=6, value=order.get_status_display())
            ws.cell(
                row=row, column=7, value=order.created_at.strftime("%d.%m.%Y %H:%M")
            )

        wb.save(response)
        return response

    export_as_excel.short_description = "Експортувати вибрані у Excel"


@admin.register(OrderItem)
class OrderItemAdmin(ExportActionModelAdmin, admin.ModelAdmin):
    resource_class = OrderItemResource
    list_display = (
        "order",
        "product",
        "product_name",
        "quantity",
        "product_price",
        "total",
    )
    list_filter = ("order__status", "created_at")
    search_fields = ("order__id", "product__name", "product_name")
    readonly_fields = ("total",)

    def total(self, obj):
        return obj.total_price()

    total.short_description = "Сума"
